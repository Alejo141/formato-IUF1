import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import calendar
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Generador Formatos SUI – Resolución 9995/2021",
    page_icon="⚡",
    layout="wide"
)

# ── Estilos ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-title {
        font-size: 1.7rem; font-weight: 700; color: #1a3a5c;
        border-bottom: 3px solid #f0a500; padding-bottom: 8px; margin-bottom: 4px;
    }
    .sub-title { font-size: 0.9rem; color: #6b7280; margin-bottom: 24px; }
    .section-header {
        font-size: 1rem; font-weight: 600; color: #1a3a5c;
        background: #eef3fa; padding: 6px 12px; border-radius: 6px;
        border-left: 4px solid #1a3a5c; margin: 16px 0 10px 0;
    }
    .stButton>button {
        background-color: #1a3a5c; color: white; border: none;
        font-weight: 600; border-radius: 6px; padding: 0.5rem 1.5rem;
    }
    .stButton>button:hover { background-color: #f0a500; color: #1a3a5c; }
    .info-box {
        background: #fffbea; border: 1px solid #f0a500;
        border-radius: 6px; padding: 10px 14px; margin-bottom: 12px;
        font-size: 0.88rem; color: #7a5500;
    }
    .success-box {
        background: #ecfdf5; border: 1px solid #10b981;
        border-radius: 6px; padding: 10px 14px; font-size: 0.88rem; color: #065f46;
    }
    .error-box {
        background: #fef2f2; border: 1px solid #ef4444;
        border-radius: 6px; padding: 10px 14px; font-size: 0.88rem; color: #7f1d1d;
    }
    .tab-desc { font-size: 0.85rem; color: #4b5563; margin-bottom: 12px; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-title">⚡ Generador Formatos SUI – SISFV ZNI</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Resolución 9995 de 2021 · Sistemas Individuales de Generación Solar Fotovoltaica – ZNI Colombia</div>', unsafe_allow_html=True)

# ── Helpers ───────────────────────────────────────────────────────────────────
MESES_ES = {
    1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
    7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"
}

def primer_dia(anio, mes):
    return date(anio, mes, 1)

def ultimo_dia(anio, mes):
    return date(anio, mes, calendar.monthrange(anio, mes)[1])

def dia5_mes_siguiente(anio, mes):
    if mes == 12:
        return date(anio + 1, 1, 5)
    return date(anio, mes + 1, 5)

def dias_mes(anio, mes):
    return calendar.monthrange(anio, mes)[1]

def fmt_fecha(d):
    return d.strftime("%d-%m-%Y")

def redondear_5dec(val):
    try: return round(float(val), 5)
    except: return val

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📅 Parámetros del formato")
    anio_sel = st.selectbox("Año", list(range(2021, date.today().year + 2)),
                            index=list(range(2021, date.today().year + 2)).index(date.today().year))
    mes_sel  = st.selectbox("Mes", list(MESES_ES.keys()),
                            format_func=lambda x: MESES_ES[x],
                            index=date.today().month - 1)
    n_dias   = dias_mes(anio_sel, mes_sel)
    fecha_ini = primer_dia(anio_sel, mes_sel)
    fecha_fin = ultimo_dia(anio_sel, mes_sel)
    fecha_exp_iuf1 = fecha_fin
    fecha_exp_f54  = dia5_mes_siguiente(anio_sel, mes_sel)
    mes_str  = str(mes_sel).zfill(2)
    anio_str = str(anio_sel)

    st.markdown("---")
    st.caption(f"**Primer día:** {fmt_fecha(fecha_ini)}")
    st.caption(f"**Último día:** {fmt_fecha(fecha_fin)}")
    st.caption(f"**Días del mes:** {n_dias}")
    st.caption(f"**FEC_EXPEDICION F54:** {fmt_fecha(fecha_exp_f54)}")
    st.markdown("---")
    st.caption("DISPOWER S.A.S. E.S.P. · ZNI Colombia")

# ── Selección de formato (tabs) ───────────────────────────────────────────────
tab_iuf1, tab_f54 = st.tabs(["📄 Formato IUF1", "📋 Formato 54"])

# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIONES COMPARTIDAS
# ═══════════════════════════════════════════════════════════════════════════════

def leer_usuarios(file_usuarios):
    df_usr = pd.read_excel(file_usuarios, sheet_name="BD_Usuarios")
    cols_req = ["NIU_SUI","COD_LOCALIDAD","Whd","LONGITUD","LATITUD","VEREDA"]
    faltantes = [c for c in cols_req if c not in df_usr.columns]
    if faltantes:
        st.markdown(f'<div class="error-box">❌ Faltan columnas en Usuarios: {faltantes}</div>', unsafe_allow_html=True)
        st.stop()
    df_usr["NIU_KEY"] = df_usr["NIU_SUI"].astype(str).str.replace("-","",regex=False).str.strip()
    return df_usr

def leer_facturacion(file_fact):
    df_fact = pd.read_excel(file_fact, sheet_name=0)
    cols_req = ["nui","nfacturasiigo","cantidad"]
    faltantes = [c for c in cols_req if c not in df_fact.columns]
    if faltantes:
        st.markdown(f'<div class="error-box">❌ Faltan columnas en Facturación: {faltantes}</div>', unsafe_allow_html=True)
        st.stop()
    df_fact["NIU_KEY"] = df_fact["nui"].astype(str).str.replace("-","",regex=False).str.strip()
    df_fact["ID_FACTURA_CLEAN"] = df_fact["nfacturasiigo"].astype(str).str.replace("-","",regex=False).str.strip()
    return df_fact

def leer_cartera(file_cartera):
    df_car = pd.read_excel(file_cartera, sheet_name="Cartera_Total_NIU")
    cols_req = ["NIU","TARIFA TOTAL"]
    faltantes = [c for c in cols_req if c not in df_car.columns]
    if faltantes:
        st.markdown(f'<div class="error-box">❌ Faltan columnas en Cartera: {faltantes}</div>', unsafe_allow_html=True)
        st.stop()
    df_car["NIU_KEY"] = df_car["NIU"].astype(str).str.replace("-","",regex=False).str.strip()
    return df_car[["NIU_KEY","TARIFA TOTAL"]].rename(columns={"TARIFA TOTAL":"CARTERA_VALOR"})

def cruce_base(df_usr, df_fact, file_cartera, mes_sel, anio_sel):
    """Cruce común: usuarios LEFT JOIN facturación LEFT JOIN cartera."""
    mes_str  = str(mes_sel).zfill(2)
    anio_str = str(anio_sel)

    df_fact_red = df_fact[["NIU_KEY","ID_FACTURA_CLEAN","cantidad"]].copy()
    df = df_usr[["NIU_KEY","NIU_SUI","COD_LOCALIDAD","Whd","LONGITUD","LATITUD","VEREDA"]].merge(
        df_fact_red, on="NIU_KEY", how="left"
    )

    sin_fact = df["ID_FACTURA_CLEAN"].isna().sum()
    en_fact  = df["ID_FACTURA_CLEAN"].notna().sum()
    if sin_fact > 0:
        st.markdown(
            f'<div class="info-box">⚠️ <strong>{sin_fact}</strong> usuario(s) sin factura este mes '
            f'(campos de facturación en 0). <strong>{en_fact}</strong> usuario(s) con factura.</div>',
            unsafe_allow_html=True
        )

    # ID FACTURA
    df["ID_FACTURA_FINAL"] = df.apply(
        lambda r: r["ID_FACTURA_CLEAN"]
        if pd.notna(r["ID_FACTURA_CLEAN"]) and r["ID_FACTURA_CLEAN"] != ""
        else f"FV{str(int(r['NIU_SUI'])).strip()}{mes_str}{anio_str}",
        axis=1
    )

    # Cartera → VALOR_MORA
    if file_cartera:
        df_car = leer_cartera(file_cartera)
        df = df.merge(df_car, on="NIU_KEY", how="left")
        df["VALOR_MORA"] = df["CARTERA_VALOR"].fillna(0.0).round(4)
        df.drop(columns=["CARTERA_VALOR"], inplace=True)
        con_mora = (df["VALOR_MORA"] > 0).sum()
        st.markdown(
            f'<div class="info-box">💳 Cartera cargada: <strong>{con_mora}</strong> NIU con saldo en VAL MORA.</div>',
            unsafe_allow_html=True
        )
    else:
        df["VALOR_MORA"] = 0.0

    # DIAS
    df["DIAS"] = df["cantidad"].fillna(0).astype(int)

    return df

def generar_xlsx_base(ws, df_out, col_widths):
    """Escribe encabezados y datos en un worksheet dado."""
    HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    DATA_FONT   = Font(name="Arial", size=9)
    ALT_FILL    = PatternFill("solid", fgColor="EEF3FA")
    BS          = Side(style="thin", color="CCCCCC")
    CB          = Border(left=BS, right=BS, bottom=BS, top=BS)

    headers = list(df_out.columns)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CB
    ws.row_dimensions[1].height = 30

    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 12)
    ws.freeze_panes = "A2"
    return headers

def generar_csv_bytes(df_out):
    buf = BytesIO()
    df_out.to_csv(buf, index=False, sep=",", encoding="utf-8-sig")
    buf.seek(0)
    return buf.getvalue()

# ═══════════════════════════════════════════════════════════════════════════════
# TAB IUF1
# ═══════════════════════════════════════════════════════════════════════════════
with tab_iuf1:
    st.markdown('<div class="tab-desc">Genera el Formato IUF1 según Resolución 9995 de 2021.</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-header">1 · Carga de archivos</div>', unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    with c1:
        st.caption("📋 **Base de Usuarios**")
        fu_iuf1 = st.file_uploader("Usuarios.xlsx", type=["xlsx"], key="u_iuf1")
    with c2:
        st.caption("🧾 **Consolidado de Facturación**")
        ff_iuf1 = st.file_uploader("Fact_MM_YYYY.xlsx", type=["xlsx"], key="f_iuf1")
    with c3:
        st.caption("💳 **Cartera por NIU** (opcional)")
        fc_iuf1 = st.file_uploader("Cartera_NIU.xlsx", type=["xlsx"], key="c_iuf1")

    if fu_iuf1 and ff_iuf1:
        try:
            df_usr  = leer_usuarios(fu_iuf1)
            df_fact = leer_facturacion(ff_iuf1)
            df      = cruce_base(df_usr, df_fact, fc_iuf1, mes_sel, anio_sel)

            # ── Campos IUF1 ──────────────────────────────────────────────────
            df["NIU"]             = df["NIU_SUI"].astype("Int64")
            df["COD LOCALIDAD"]   = df["COD_LOCALIDAD"].astype(str)
            df["dane"]            = df["COD LOCALIDAD"].str[:5]
            df["Whd_col"]         = df["Whd"]
            df["ALTITUD"]         = 0
            df["LONGITUD"]        = df["LONGITUD"].apply(redondear_5dec)
            df["LATITUD"]         = df["LATITUD"].apply(redondear_5dec)
            df["ID FACTURA"]      = df["ID_FACTURA_FINAL"]
            df["TIPO CORRI SALIDA"] = 1
            df["DIAS PRES MES"]   = df["DIAS"]
            df["ENERGIA GEN MES"] = (df["Whd"] * df["DIAS PRES MES"] / 1000).round(3).fillna(0)
            df["disp promedio"]   = (df["DIAS PRES MES"] / n_dias).round(4)
            df["IUC"]  = 0.0000
            df["MP"]   = 0.0000
            df["GIO"]  = 0.0000
            df["PE"]   = 0.0000
            df["COR"]  = ""
            df["GAOM"] = ""
            df["DIRECCION"]       = df["VEREDA"]
            df["FECH EXP FACT"]   = fmt_fecha(fecha_exp_iuf1)
            df["FECH INI PERIO"]  = fmt_fecha(fecha_ini)
            df["DIAS FACT"]       = df["DIAS PRES MES"]
            df["ESTRATO"]         = 1
            df["TIPO LECT"]       = 3
            df["FACT CONSUMO"]    = df["ID_FACTURA_CLEAN"].apply(lambda x: "" if pd.notna(x) and x != "" else 0)
            df["VAL REFACT"]      = 0.0000
            df["VAL MORA"]        = df["VALOR_MORA"].round(4)
            df["INT MORA"]        = 0.0000
            df["VAL SUBS"]        = df["ID_FACTURA_CLEAN"].apply(lambda x: "" if pd.notna(x) and x != "" else 0)
            df["PORCE SUBS"]      = df["ID_FACTURA_CLEAN"].apply(lambda x: "" if pd.notna(x) and x != "" else 0)
            df["TARIFA"]          = df["ID_FACTURA_CLEAN"].apply(lambda x: "" if pd.notna(x) and x != "" else 0)
            df["VAL TOTAL FACT"]  = ""

            COLS_IUF1 = [
                "NIU","COD LOCALIDAD","dane","Whd_col","ALTITUD","LONGITUD","LATITUD",
                "ID FACTURA","ENERGIA GEN MES","TIPO CORRI SALIDA","DIAS PRES MES",
                "disp promedio","IUC","MP","COR","GIO","PE","GAOM","DIRECCION",
                "FECH EXP FACT","FECH INI PERIO","DIAS FACT","ESTRATO","TIPO LECT",
                "FACT CONSUMO","VAL REFACT","VAL MORA","INT MORA","VAL SUBS",
                "PORCE SUBS","TARIFA","VAL TOTAL FACT"
            ]
            df_out = df[COLS_IUF1].copy()
            df_out = df_out.rename(columns={"Whd_col":"Whd"})

            # ── Vista previa ─────────────────────────────────────────────────
            st.markdown('<div class="section-header">2 · Vista previa</div>', unsafe_allow_html=True)
            st.caption(f"**{len(df_out):,}** registros · {MESES_ES[mes_sel]} {anio_sel}")
            st.dataframe(df_out.head(50), use_container_width=True, height=260)

            with st.expander("📊 Resumen"):
                m1,m2,m3,m4 = st.columns(4)
                m1.metric("Registros IUF1", len(df_out))
                m2.metric("Con factura", int(df["ID_FACTURA_CLEAN"].notna().sum()))
                m3.metric("Sin factura", int(df["ID_FACTURA_CLEAN"].isna().sum()))
                m4.metric("Días del mes", n_dias)

            # ── Exportar ─────────────────────────────────────────────────────
            st.markdown('<div class="section-header">3 · Descargar</div>', unsafe_allow_html=True)

            def build_xlsx_iuf1(df_out):
                wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Hoja1"
                COL_W = {
                    "NIU":14,"COD LOCALIDAD":17,"dane":8,"Whd":6,"ALTITUD":7,
                    "LONGITUD":12,"LATITUD":12,"ID FACTURA":16,"ENERGIA GEN MES":14,
                    "TIPO CORRI SALIDA":8,"DIAS PRES MES":10,"disp promedio":11,
                    "IUC":9,"MP":9,"COR":9,"GIO":9,"PE":9,"GAOM":9,
                    "DIRECCION":18,"FECH EXP FACT":13,"FECH INI PERIO":13,
                    "DIAS FACT":9,"ESTRATO":8,"TIPO LECT":9,"FACT CONSUMO":13,
                    "VAL REFACT":11,"VAL MORA":12,"INT MORA":10,"VAL SUBS":12,
                    "PORCE SUBS":10,"TARIFA":12,"VAL TOTAL FACT":13,
                }
                headers = generar_xlsx_base(ws, df_out, COL_W)
                col_idx = {h: i+1 for i, h in enumerate(headers)}
                BS = Side(style="thin", color="CCCCCC")
                CB = Border(left=BS, right=BS, bottom=BS, top=BS)
                ALT = PatternFill("solid", fgColor="EEF3FA")
                DF  = Font(name="Arial", size=9)
                COLS_4DEC = {"IUC","MP","GIO","PE","VAL REFACT","VAL MORA","INT MORA"}
                COLS_5DEC = {"LONGITUD","LATITUD"}
                for ri, row in enumerate(df_out.itertuples(index=False), 2):
                    fill = ALT if ri % 2 == 0 else None
                    for ci, (h, val) in enumerate(zip(headers, list(row)), 1):
                        cell = ws.cell(row=ri, column=ci)
                        cell.font = DF; cell.border = CB
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        if fill: cell.fill = fill
                        col_dp  = get_column_letter(col_idx["DIAS PRES MES"])
                        col_fc  = get_column_letter(col_idx["FACT CONSUMO"])
                        col_vs  = get_column_letter(col_idx["VAL SUBS"])
                        if h == "disp promedio":
                            cell.value = f"={col_dp}{ri}/{n_dias}"; cell.number_format = "#,##0.0000"
                        elif h == "PORCE SUBS":
                            cell.value = f'=IFERROR({col_vs}{ri}/{col_fc}{ri},"")'; cell.number_format = "#,##0.0000"
                        elif h == "VAL TOTAL FACT":
                            cell.value = f'=IFERROR({col_fc}{ri},"")'; cell.number_format = "#,##0.0000"
                        elif h in COLS_5DEC:
                            cell.value = float(val) if val != "" else ""; cell.number_format = "#,##0.00000"
                        elif h in COLS_4DEC:
                            cell.value = float(val) if val != "" else 0.0; cell.number_format = "#,##0.0000"
                        elif h == "ENERGIA GEN MES":
                            cell.value = float(val) if val != "" else ""; cell.number_format = "#,##0.000"
                        elif h == "NIU":
                            cell.value = int(val) if pd.notna(val) else ""
                            cell.number_format = "0"
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        elif h in {"COD LOCALIDAD","dane"}:
                            cell.value = str(val) if val != "" else ""
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        else:
                            cell.value = val if val != "" else (None if isinstance(val, str) else val)
                buf = BytesIO(); wb.save(buf); buf.seek(0)
                return buf.getvalue()

            nombre_iuf1 = f"IUF1_{mes_str}_{anio_str}"
            d1, d2 = st.columns(2)
            with d1:
                st.download_button(f"⬇️ {nombre_iuf1}.xlsx", build_xlsx_iuf1(df_out),
                    f"{nombre_iuf1}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                st.caption("Incluye fórmulas Excel para disp promedio, PORCE SUBS y VAL TOTAL FACT")
            with d2:
                st.download_button(f"⬇️ {nombre_iuf1}.csv", generar_csv_bytes(df_out),
                    f"{nombre_iuf1}.csv", "text/csv", use_container_width=True)
                st.caption("CSV separado por comas · UTF-8")

            st.markdown(f'<div class="success-box">✅ IUF1 generado: <strong>{MESES_ES[mes_sel]} {anio_sel}</strong> · <strong>{len(df_out):,} registros</strong>.</div>', unsafe_allow_html=True)

        except Exception as e:
            st.markdown(f'<div class="error-box">❌ Error: {e}</div>', unsafe_allow_html=True)
            st.exception(e)
    else:
        st.markdown('<div class="info-box">👆 Carga los archivos de <strong>Usuarios</strong> y <strong>Facturación</strong> para generar el IUF1.</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB FORMATO 54
# ═══════════════════════════════════════════════════════════════════════════════
with tab_f54:
    st.markdown('<div class="tab-desc">Genera el Formato 54 (ZNISISFV) según Resolución 9995 de 2021.</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-header">1 · Carga de archivos</div>', unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.caption("📋 **Base de Usuarios**")
        fu_f54 = st.file_uploader("Usuarios.xlsx", type=["xlsx"], key="u_f54")
    with c2:
        st.caption("🧾 **Consolidado de Facturación**")
        ff_f54 = st.file_uploader("Fact_MM_YYYY.xlsx", type=["xlsx"], key="f_f54")
    with c3:
        st.caption("💳 **Cartera por NIU** (opcional)")
        fc_f54 = st.file_uploader("Cartera_NIU.xlsx", type=["xlsx"], key="c_f54")
    with c4:
        st.caption("📊 **Consolidado de Cálculos**")
        fcon_f54 = st.file_uploader("Consolidado_YYYY_MM.xlsx", type=["xlsx"], key="con_f54")

    if fu_f54 and ff_f54 and fcon_f54:
        try:
            df_usr  = leer_usuarios(fu_f54)
            df_fact = leer_facturacion(ff_f54)
            df      = cruce_base(df_usr, df_fact, fc_f54, mes_sel, anio_sel)

            # ── Leer Consolidado de Cálculos ─────────────────────────────────
            df_con = pd.read_excel(fcon_f54, sheet_name="Consolidado")
            cols_con = ["llave","AMGCnu_m","AMGCvi_m","AMGCau_m","AMGCnf_m","AMGCro_m",
                        "Subsidio_dia","tarifa_dia","fact_dia"]
            faltantes_con = [c for c in cols_con if c not in df_con.columns]
            if faltantes_con:
                st.markdown(f'<div class="error-box">❌ Faltan columnas en Consolidado: {faltantes_con}</div>', unsafe_allow_html=True)
                st.stop()
            df_con["llave"] = df_con["llave"].astype(str).str.strip()
            df_con = df_con[cols_con].copy()

            # ── Construir campos Formato 54 ───────────────────────────────────

            # Campos base
            df["Whd_col"]       = df["Whd"]
            df["dane"]          = df["COD_LOCALIDAD"].astype(str).str[:5]
            df["llave"]         = df["dane"] + df["Whd"].astype(str)
            df["NIU"]           = df["NIU_SUI"].astype("Int64")
            df["COD_LOC"]       = df["COD_LOCALIDAD"].astype(str)
            df["ID_FACTURA"]    = df["ID_FACTURA_FINAL"]
            df["DIAS_FAC"]      = df["DIAS"]

            # Cruce con consolidado por llave
            df = df.merge(df_con, on="llave", how="left")

            sin_con = df["AMGCnu_m"].isna().sum()
            if sin_con > 0:
                st.markdown(
                    f'<div class="info-box">⚠️ <strong>{sin_con}</strong> registro(s) sin match en el Consolidado de Cálculos.</div>',
                    unsafe_allow_html=True
                )

            # Campos calculados
            df["FEC_EXPEDICION"]     = fmt_fecha(fecha_exp_f54)
            df["ESQ_FACTURACION"]    = 1
            df["FEC_INICIO_PERIODO"] = fmt_fecha(fecha_ini)
            df["FEC_FIN_PERIODO"]    = fmt_fecha(fecha_fin)
            df["DISP_PERIODO"]       = df["DIAS_FAC"].astype(float).round(2)
            df["disp"]               = (df["DIAS_FAC"] / n_dias).round(2)
            df["CONSUMO_ENERGIA"]    = (df["Whd"] * df["DIAS_FAC"] / 1000).round(2).fillna(0)
            df["DIAS_PRESTACION"]    = df["DIAS_FAC"]
            df["CARGO_INVERSION"]    = 0.000
            df["TIPO_LECTURA"]       = 2
            df["FACTOR_CONSUMO"]     = df["DIAS_FAC"].apply(lambda x: 1 if x > 0 else 0)
            df["CANT_MINIMA"]        = df["CONSUMO_ENERGIA"]

            # Desde consolidado × DIAS_FACTURADOS
            df["VALOR_SUBSIDIO"]     = (df["Subsidio_dia"] * df["DIAS_FAC"]).round(4).fillna(0)
            df["VALOR_TARIFA"]       = (df["tarifa_dia"]   * df["DIAS_FAC"]).round(4).fillna(0)
            df["FACT_CONSUMO_F54"]   = (df["fact_dia"]     * df["DIAS_FAC"]).round(4).fillna(0)

            # AMGC campos (3 decimales)
            for src, dst in [("AMGCnu_m","AMGC_NUM_USUARIOS"),("AMGCvi_m","AMGC_VALOR_INVERSION"),
                              ("AMGCau_m","AMGC_ATENCION"),("AMGCnf_m","AMGC_NIVEL_FAC"),
                              ("AMGCro_m","CARGO_REMUNERACION")]:
                df[dst] = df[src].round(3).fillna(0)

            df["VALOR_REFAC"]        = 0.0000
            df["VALOR_MORA_F54"]     = df["VALOR_MORA"].round(4)
            df["INTERES_MORA"]       = 0.0000
            df["VALOR_TOTAL_FAC"]    = (df["VALOR_TARIFA"] + df["VALOR_MORA_F54"]).round(4)
            df["VALOR_CARTERA_REC"]  = 0.0000

            COLS_F54 = [
                "Whd_col","dane","llave","NIU","COD_LOC","ID_FACTURA",
                "FEC_EXPEDICION","ESQ_FACTURACION","FEC_INICIO_PERIODO","FEC_FIN_PERIODO",
                "DIAS_FAC","DISP_PERIODO","disp","CONSUMO_ENERGIA","DIAS_PRESTACION",
                "CARGO_INVERSION","AMGC_NUM_USUARIOS","AMGC_VALOR_INVERSION","AMGC_ATENCION",
                "AMGC_NIVEL_FAC","CARGO_REMUNERACION","TIPO_LECTURA","FACTOR_CONSUMO",
                "CANT_MINIMA","VALOR_SUBSIDIO","VALOR_TARIFA","FACT_CONSUMO_F54",
                "VALOR_REFAC","VALOR_MORA_F54","INTERES_MORA","VALOR_TOTAL_FAC","VALOR_CARTERA_REC"
            ]
            COLS_RENAME = {
                "Whd_col":"Whd","COD_LOC":"COD_LOCALIDAD","DIAS_FAC":"DIAS_FACTURADOS",
                "FACT_CONSUMO_F54":"FACT_CONSUMO","VALOR_MORA_F54":"VALOR_MORA"
            }
            df_out54 = df[COLS_F54].rename(columns=COLS_RENAME).copy()

            # ── Vista previa ─────────────────────────────────────────────────
            st.markdown('<div class="section-header">2 · Vista previa</div>', unsafe_allow_html=True)
            st.caption(f"**{len(df_out54):,}** registros · {MESES_ES[mes_sel]} {anio_sel}")
            st.dataframe(df_out54.head(50), use_container_width=True, height=260)

            with st.expander("📊 Resumen"):
                m1,m2,m3,m4 = st.columns(4)
                m1.metric("Registros F54", len(df_out54))
                m2.metric("Con factura", int(df["ID_FACTURA_CLEAN"].notna().sum()))
                m3.metric("Sin factura", int(df["ID_FACTURA_CLEAN"].isna().sum()))
                m4.metric("Días del mes", n_dias)

            # ── Exportar ─────────────────────────────────────────────────────
            st.markdown('<div class="section-header">3 · Descargar</div>', unsafe_allow_html=True)

            def build_xlsx_f54(df_out54):
                wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Formato54"
                COL_W = {
                    "Whd":6,"dane":8,"llave":12,"NIU":14,"COD_LOCALIDAD":17,
                    "ID_FACTURA":18,"FEC_EXPEDICION":13,"ESQ_FACTURACION":8,
                    "FEC_INICIO_PERIODO":14,"FEC_FIN_PERIODO":13,"DIAS_FACTURADOS":10,
                    "DISP_PERIODO":10,"disp":8,"CONSUMO_ENERGIA":13,"DIAS_PRESTACION":10,
                    "CARGO_INVERSION":10,"AMGC_NUM_USUARIOS":14,"AMGC_VALOR_INVERSION":16,
                    "AMGC_ATENCION":12,"AMGC_NIVEL_FAC":12,"CARGO_REMUNERACION":14,
                    "TIPO_LECTURA":9,"FACTOR_CONSUMO":10,"CANT_MINIMA":11,
                    "VALOR_SUBSIDIO":13,"VALOR_TARIFA":12,"FACT_CONSUMO":13,
                    "VALOR_REFAC":11,"VALOR_MORA":12,"INTERES_MORA":11,
                    "VALOR_TOTAL_FAC":13,"VALOR_CARTERA_REC":14,
                }
                headers = generar_xlsx_base(ws, df_out54, COL_W)
                BS = Side(style="thin", color="CCCCCC")
                CB = Border(left=BS, right=BS, bottom=BS, top=BS)
                ALT = PatternFill("solid", fgColor="EEF3FA")
                DF  = Font(name="Arial", size=9)

                COLS_4DEC = {"VALOR_SUBSIDIO","VALOR_TARIFA","FACT_CONSUMO","VALOR_REFAC",
                             "VALOR_MORA","INTERES_MORA","VALOR_TOTAL_FAC","VALOR_CARTERA_REC"}
                COLS_3DEC = {"CARGO_INVERSION","AMGC_NUM_USUARIOS","AMGC_VALOR_INVERSION",
                             "AMGC_ATENCION","AMGC_NIVEL_FAC","CARGO_REMUNERACION"}
                COLS_2DEC = {"DISP_PERIODO","CONSUMO_ENERGIA","disp"}
                COLS_INT  = {"NIU","ESQ_FACTURACION","DIAS_FACTURADOS","DIAS_PRESTACION",
                             "TIPO_LECTURA","FACTOR_CONSUMO"}
                COLS_STR  = {"dane","llave","COD_LOCALIDAD","ID_FACTURA",
                             "FEC_EXPEDICION","FEC_INICIO_PERIODO","FEC_FIN_PERIODO"}

                for ri, row in enumerate(df_out54.itertuples(index=False), 2):
                    fill = ALT if ri % 2 == 0 else None
                    for ci, (h, val) in enumerate(zip(headers, list(row)), 1):
                        cell = ws.cell(row=ri, column=ci)
                        cell.font = DF; cell.border = CB
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        if fill: cell.fill = fill
                        if h in COLS_4DEC:
                            cell.value = float(val) if pd.notna(val) else 0.0
                            cell.number_format = "#,##0.0000"
                        elif h in COLS_3DEC:
                            cell.value = float(val) if pd.notna(val) else 0.0
                            cell.number_format = "#,##0.000"
                        elif h in COLS_2DEC:
                            cell.value = float(val) if pd.notna(val) else 0.0
                            cell.number_format = "#,##0.00"
                        elif h in COLS_INT:
                            cell.value = int(val) if pd.notna(val) else 0
                            cell.number_format = "0"
                        elif h in COLS_STR:
                            cell.value = str(val) if pd.notna(val) else ""
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        else:
                            cell.value = float(val) if pd.notna(val) else 0.0
                            cell.number_format = "#,##0.000"
                buf = BytesIO(); wb.save(buf); buf.seek(0)
                return buf.getvalue()

            nombre_f54 = f"Formato54_{mes_str}_{anio_str}"
            d1, d2 = st.columns(2)
            with d1:
                st.download_button(f"⬇️ {nombre_f54}.xlsx", build_xlsx_f54(df_out54),
                    f"{nombre_f54}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            with d2:
                st.download_button(f"⬇️ {nombre_f54}.csv", generar_csv_bytes(df_out54),
                    f"{nombre_f54}.csv", "text/csv", use_container_width=True)

            st.markdown(f'<div class="success-box">✅ Formato 54 generado: <strong>{MESES_ES[mes_sel]} {anio_sel}</strong> · <strong>{len(df_out54):,} registros</strong>.</div>', unsafe_allow_html=True)

        except Exception as e:
            st.markdown(f'<div class="error-box">❌ Error: {e}</div>', unsafe_allow_html=True)
            st.exception(e)
    else:
        st.markdown('<div class="info-box">👆 Carga los archivos de <strong>Usuarios</strong>, <strong>Facturación</strong> y <strong>Consolidado de Cálculos</strong> para generar el Formato 54.</div>', unsafe_allow_html=True)
